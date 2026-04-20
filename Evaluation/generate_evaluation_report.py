"""
Mudra Evaluation Report Generator
===================================
Reads the human-feedback Excel and produces a clean, per-section workbook.

Layout
------
One tab per evaluation section (Present Tense, WH-Questions, Past Tense,
Imperative, Negation, WH-Q+Past Tense, Names/Fingerspelling, Numbers,
Sentence Clips, Session Feedback).

Within each tab:
  - Fixed sentences   : one block per sentence showing all evaluators
  - User-input sents  : each evaluator's unique sentence shown individually
  - Summary row (fixed only): avg accuracy, avg human-likeness, top issues
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────
#  Input / Output paths
# ─────────────────────────────────────────────────────────────────
INPUT_FILE  = "Mudra evaluation - I (1).xlsx"
OUTPUT_FILE = "Mudra_Evaluation_By_Section_v2.xlsx"

# ─────────────────────────────────────────────────────────────────
#  Style helpers
# ─────────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def _font(bold=False, italic=False, color="000000", size=10):
    return Font(bold=bold, italic=italic, color=color, size=size)

F_SECTION   = _fill("1F3864")   # dark navy  – section title
F_FIXED     = _fill("2E75B6")   # mid blue   – fixed sentence header
F_INPUT     = _fill("375623")   # dark green – user-input sentence header
F_COL_HDR   = _fill("BDD7EE")   # pale blue  – column headers
F_SUMMARY   = _fill("E2EFDA")   # pale green – summary row
F_PLACEHOLDER = _fill("FFF2CC") # pale yellow – fixed-sentence cell
F_ODD       = _fill("FFFFFF")
F_EVEN      = _fill("F2F2F2")

FONT_WHITE_BOLD  = _font(bold=True,  color="FFFFFF", size=10)
FONT_BOLD        = _font(bold=True,  size=10)
FONT_ITALIC_GREY = _font(italic=True, color="888888", size=9)
FONT_NORMAL      = _font(size=10)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ─────────────────────────────────────────────────────────────────
#  Output column definitions (metadata cols + feedback cols)
# ─────────────────────────────────────────────────────────────────
# Standard 8-feedback-col sections
STD_COLS      = ["Evaluator", "Sentence",
                 "Watch Times", "Accuracy (1–5)", "Word–Sign Mapping",
                 "Facial Expr. & NMM", "Human-likeness (1–5)",
                 "Issues Observed", "Special Check", "Comments"]

# Names / Numbers 11-feedback-col sections (Part III sents 1-11)
DETAIL_COLS   = ["Evaluator", "Sentence",
                 "Watch Times", "Accuracy (1–5)", "Word–Sign Mapping",
                 "Facial Expr. & NMM", "Human-likeness (1–5)",
                 "Issues Observed", "Num / Special Check",
                 "Transition Check", "Fingerspelling Accuracy",
                 "Name Identification", "Comments"]

# Sentence-Clips 8-feedback-col (Part III sents 12-13)
CLIPS_COLS    = ["Evaluator", "Sentence",
                 "Watch Times", "Accuracy (1–5)", "Word–Sign Mapping",
                 "Facial Expr. & NMM", "Human-likeness (1–5)",
                 "Issues Observed", "Video Transition Quality", "Comments"]

# ─────────────────────────────────────────────────────────────────
#  Section definitions
#  Each sentence entry:
#    num        : sentence number within this section
#    text       : None  (fixed → pulled from form; user-input → from data)
#    is_input   : True  = user typed sentence is in the data row
#    input_col  : 0-based column index of the sentence text (is_input only)
#    data_start : 0-based column index of first feedback column
# ─────────────────────────────────────────────────────────────────
SECTIONS = [
    # ── Sheet 1 (Part I) ────────────────────────────────────────
    {
        "sheet_key": "Form responses 1",
        "tab_name":  "1. Present Tense",
        "title":     "Present Tense Evaluation  (Form Part I)",
        "col_names": STD_COLS,
        "n_cols":    8,
        "sentences": [
            # Fixed (text from Google Form – please fill the yellow cells)
            {"num": 1, "text": None, "is_input": False, "data_start": 8},
            {"num": 2, "text": None, "is_input": False, "data_start": 16},
            {"num": 3, "text": None, "is_input": False, "data_start": 24},
            {"num": 4, "text": None, "is_input": False, "data_start": 32},
            # User-input (unique per evaluator)
            {"num": 5, "text": None, "is_input": True, "input_col": 40, "data_start": 41},
            {"num": 6, "text": None, "is_input": True, "input_col": 49, "data_start": 50},
            {"num": 7, "text": None, "is_input": True, "input_col": 58, "data_start": 59},
        ],
    },
    {
        "sheet_key": "Form responses 1",
        "tab_name":  "2. WH-Questions",
        "title":     "WH-Questions (Interrogative) Evaluation  (Form Part I)",
        "col_names": STD_COLS,
        "n_cols":    8,
        "sentences": [
            {"num": 1, "text": None, "is_input": False, "data_start": 67},
            {"num": 2, "text": None, "is_input": False, "data_start": 75},
            {"num": 3, "text": None, "is_input": False, "data_start": 83},
            {"num": 4, "text": None, "is_input": False, "data_start": 91},
            {"num": 5, "text": None, "is_input": True, "input_col": 99,  "data_start": 100},
            {"num": 6, "text": None, "is_input": True, "input_col": 108, "data_start": 109},
        ],
    },
    # ── Sheet 2 (Part II) ───────────────────────────────────────
    {
        "sheet_key": "Form responses 2",
        "tab_name":  "3. Past Tense",
        "title":     "Past Tense Evaluation  (Form Part II)",
        "col_names": STD_COLS,
        "n_cols":    8,
        "sentences": [
            {"num": 1, "text": None, "is_input": False, "data_start":  2},
            {"num": 2, "text": None, "is_input": False, "data_start": 10},
            {"num": 3, "text": None, "is_input": False, "data_start": 18},
            {"num": 4, "text": None, "is_input": False, "data_start": 26},
            {"num": 5, "text": None, "is_input": True,  "input_col": 34, "data_start": 35},
            {"num": 6, "text": None, "is_input": True,  "input_col": 43, "data_start": 44},
        ],
    },
    {
        "sheet_key": "Form responses 2",
        "tab_name":  "4. Imperative",
        "title":     "Imperative (Command) Sentences Evaluation  (Form Part II)",
        "col_names": STD_COLS,
        "n_cols":    8,
        "sentences": [
            {"num": 1, "text": None, "is_input": False, "data_start": 52},
            {"num": 2, "text": None, "is_input": False, "data_start": 60},
            {"num": 3, "text": None, "is_input": False, "data_start": 68},
            {"num": 4, "text": None, "is_input": False, "data_start": 76},
            {"num": 5, "text": None, "is_input": True,  "input_col": 84, "data_start": 85},
            {"num": 6, "text": None, "is_input": True,  "input_col": 93, "data_start": 94},
        ],
    },
    {
        "sheet_key": "Form responses 2",
        "tab_name":  "5. Negation",
        "title":     "Negation Sentences Evaluation  (Form Part II)",
        "col_names": STD_COLS,
        "n_cols":    8,
        "sentences": [
            {"num": 1, "text": None, "is_input": False, "data_start": 102},
            {"num": 2, "text": None, "is_input": False, "data_start": 110},
            {"num": 3, "text": None, "is_input": True,  "input_col": 118, "data_start": 119},
            {"num": 4, "text": None, "is_input": True,  "input_col": 127, "data_start": 128},
            {"num": 5, "text": None, "is_input": True,  "input_col": 136, "data_start": 137},
        ],
    },
    {
        "sheet_key": "Form responses 2",
        "tab_name":  "6. WH-Q + Past Tense",
        "title":     "WH-Question + Past Tense Evaluation  (Form Part II)",
        "col_names": STD_COLS,
        "n_cols":    8,
        "sentences": [
            {"num": 1, "text": None, "is_input": False, "data_start": 145},
            {"num": 2, "text": None, "is_input": False, "data_start": 153},
            {"num": 3, "text": None, "is_input": False, "data_start": 161},
        ],
    },
    # ── Sheet 3 (Part III) ──────────────────────────────────────
    {
        "sheet_key": "Form responses 3",
        "tab_name":  "7. Names & Fingerspell",
        "title":     "Names & Fingerspelling Evaluation  (Form Part III)",
        "col_names": DETAIL_COLS,
        "n_cols":    11,
        "sentences": [
            {"num": 1, "text": None, "is_input": False, "data_start":  2},
            {"num": 2, "text": None, "is_input": False, "data_start": 13},
            {"num": 3, "text": None, "is_input": False, "data_start": 24},
            {"num": 4, "text": None, "is_input": False, "data_start": 35},
            {"num": 5, "text": None, "is_input": False, "data_start": 46},
        ],
    },
    {
        "sheet_key": "Form responses 3",
        "tab_name":  "8. Numbers",
        "title":     "Number Recognition Evaluation  (Form Part III)",
        "col_names": DETAIL_COLS,
        "n_cols":    11,
        "sentences": [
            {"num": 1, "text": None, "is_input": False, "data_start":  57},
            {"num": 2, "text": None, "is_input": False, "data_start":  68},
            {"num": 3, "text": None, "is_input": False, "data_start":  79},
            {"num": 4, "text": None, "is_input": False, "data_start":  90},
            {"num": 5, "text": None, "is_input": False, "data_start": 101},
            {"num": 6, "text": None, "is_input": False, "data_start": 112},
        ],
    },
    {
        "sheet_key": "Form responses 3",
        "tab_name":  "9. Sentence Clips",
        "title":     "Composite Sentence Clips Evaluation  (Form Part III)",
        "col_names": CLIPS_COLS,
        "n_cols":    8,
        "sentences": [
            {"num": 1, "text": None, "is_input": True, "input_col": 123, "data_start": 124},
            {"num": 2, "text": None, "is_input": True, "input_col": 132, "data_start": 133},
        ],
    },
]

# ─────────────────────────────────────────────────────────────────
#  Utilities
# ─────────────────────────────────────────────────────────────────
def safe_mean(values):
    nums = [v for v in values if isinstance(v, (int, float)) and v is not None]
    return round(sum(nums) / len(nums), 2) if nums else None


def top_issues_str(evaluations, issues_fb_index=5):
    counter = {}
    for ev in evaluations:
        val = ev["feedback"][issues_fb_index]
        if val:
            for part in str(val).split(","):
                p = part.strip()
                if p and p.lower() not in ("none", ""):
                    counter[p] = counter.get(p, 0) + 1
    if not counter:
        return "None reported"
    sorted_issues = sorted(counter.items(), key=lambda x: -x[1])[:4]
    return ";  ".join(f"{k} (×{v})" for k, v in sorted_issues)


def write_cell(ws, row, col, value, fill=None, font=None, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:  cell.fill  = fill
    if font:  cell.font  = font
    if align: cell.alignment = align
    return cell


def safe_merge(ws, r1, c1, r2, c2):
    if r1 == r2 and c1 == c2:
        return
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


# ─────────────────────────────────────────────────────────────────
#  Tab builder
# ─────────────────────────────────────────────────────────────────
def build_section_tab(wb_out, section, ws_in):
    ws = wb_out.create_sheet(title=section["tab_name"])
    col_names = section["col_names"]
    max_c     = len(col_names)
    n_fb      = section["n_cols"]
    row       = 1

    # Section title
    write_cell(ws, row, 1, section["title"],
               fill=F_SECTION, font=FONT_WHITE_BOLD, align=ALIGN_LEFT)
    safe_merge(ws, row, 1, row, max_c)
    ws.row_dimensions[row].height = 22
    row += 2

    # ── collect all data rows from input sheet ──
    all_data = [r for r in ws_in.iter_rows(min_row=2, values_only=True)
                if r[1]]  # skip blank rows

    # ── iterate sentences ──
    for sent_def in section["sentences"]:
        snum       = sent_def["num"]
        data_start = sent_def["data_start"]
        is_input   = sent_def["is_input"]

        # collect evaluator entries for this sentence
        evaluations = []
        for dr in all_data:
            if is_input:
                s_text = dr[sent_def["input_col"]]
                if not s_text:
                    s_text = "(not provided)"
            else:
                s_text = None   # placeholder – filled below per-cell

            feedback = list(dr[data_start: data_start + n_fb])
            evaluations.append({
                "user":     dr[1],
                "sentence": s_text,
                "feedback": feedback,
            })

        # ── sentence header ──
        if is_input:
            hdr_text = f"  Sentence {snum}  ▸  User-Input  (each evaluator typed their own sentence)"
            hdr_fill = F_INPUT
        else:
            hdr_text = f"  Sentence {snum}  ▸  Fixed  (same sentence shown to all evaluators – fill yellow cell from form)"
            hdr_fill = F_FIXED

        write_cell(ws, row, 1, hdr_text,
                   fill=hdr_fill, font=FONT_WHITE_BOLD, align=ALIGN_LEFT)
        safe_merge(ws, row, 1, row, max_c)
        ws.row_dimensions[row].height = 18
        row += 1

        # ── column headers ──
        for c, cn in enumerate(col_names, 1):
            write_cell(ws, row, c, cn, fill=F_COL_HDR, font=FONT_BOLD, align=ALIGN_CENTER)
        ws.row_dimensions[row].height = 30
        row += 1

        # ── data rows ──
        for idx, ev in enumerate(evaluations):
            f_row = F_ODD if idx % 2 == 0 else F_EVEN

            # Evaluator cell
            write_cell(ws, row, 1, ev["user"], fill=f_row, font=FONT_BOLD, align=ALIGN_CENTER)

            # Sentence text cell
            if is_input:
                write_cell(ws, row, 2, ev["sentence"], fill=f_row,
                           font=FONT_NORMAL, align=ALIGN_LEFT)
            else:
                # Yellow placeholder for fixed sentences
                write_cell(ws, row, 2, "[Fixed – enter sentence from form]",
                           fill=F_PLACEHOLDER, font=FONT_ITALIC_GREY, align=ALIGN_LEFT)

            # Feedback cells
            for fi, fval in enumerate(ev["feedback"]):
                col_pos = fi + 3
                if isinstance(fval, (int, float)) and fval is not None:
                    write_cell(ws, row, col_pos, fval, fill=f_row,
                               font=FONT_BOLD, align=ALIGN_CENTER)
                else:
                    write_cell(ws, row, col_pos, fval, fill=f_row,
                               font=FONT_NORMAL, align=ALIGN_LEFT)

            ws.row_dimensions[row].height = 45
            row += 1

        # ── summary row (fixed sentences only) ──
        if not is_input and evaluations:
            acc_vals = [ev["feedback"][1] for ev in evaluations]
            hum_vals = [ev["feedback"][4] for ev in evaluations]
            avg_acc  = safe_mean(acc_vals)
            avg_hum  = safe_mean(hum_vals)
            issues_s = top_issues_str(evaluations, issues_fb_index=5)

            write_cell(ws, row, 1, "▶ SUMMARY", fill=F_SUMMARY,
                       font=FONT_BOLD, align=ALIGN_CENTER)

            acc_str = f"Avg Accuracy: {avg_acc if avg_acc is not None else 'N/A'} / 5   |   " \
                      f"Avg Human-likeness: {avg_hum if avg_hum is not None else 'N/A'} / 5"
            write_cell(ws, row, 2, acc_str, fill=F_SUMMARY,
                       font=FONT_BOLD, align=ALIGN_LEFT)
            safe_merge(ws, row, 2, row, 4)

            write_cell(ws, row, 5, f"Top Issues:  {issues_s}", fill=F_SUMMARY,
                       font=FONT_NORMAL, align=ALIGN_LEFT)
            safe_merge(ws, row, 5, row, max_c)

            ws.row_dimensions[row].height = 20
            row += 1

        row += 1   # blank separator between sentences

    # ── freeze top rows and set column widths ──
    ws.freeze_panes = "A3"

    ws.column_dimensions["A"].width = 13   # Evaluator
    ws.column_dimensions["B"].width = 44   # Sentence
    width_map = {
        "Watch Times": 34,   "Accuracy (1–5)": 13, "Human-likeness (1–5)": 14,
        "Issues Observed": 34, "Comments": 38, "Word–Sign Mapping": 28,
        "Facial Expr. & NMM": 28, "Special Check": 30,
        "Num / Special Check": 28, "Transition Check": 28,
        "Fingerspelling Accuracy": 28, "Name Identification": 28,
        "Video Transition Quality": 28,
    }
    for ci, cn in enumerate(col_names[2:], 3):
        ws.column_dimensions[get_column_letter(ci)].width = width_map.get(cn, 22)


# ─────────────────────────────────────────────────────────────────
#  Participant Details tab  (Form Part I – demographic columns)
# ─────────────────────────────────────────────────────────────────
def build_participant_details_tab(wb_out, ws1):
    ws = wb_out.create_sheet(title="0. Participant Details")
    row = 1

    title = "Participant Details  (Form Part I – demographic & background information)"
    write_cell(ws, row, 1, title, fill=F_SECTION,
               font=FONT_WHITE_BOLD, align=ALIGN_LEFT)
    safe_merge(ws, row, 1, row, 8)
    ws.row_dimensions[row].height = 22
    row += 2

    # Column headers  (0-based source indices in parentheses)
    headers = [
        "#",
        "Timestamp",
        "User ID",
        "Age",
        "Participant Type",
        "Hearing Status",
        "Sign Proficiency",
        "Years of Signing",
        "Studied SL in School?",
    ]
    for c, h in enumerate(headers, 1):
        write_cell(ws, row, c, h, fill=F_COL_HDR, font=FONT_BOLD, align=ALIGN_CENTER)
    ws.row_dimensions[row].height = 30
    row += 1

    # Data rows
    # Source col indices (0-based):  0=Timestamp, 1=UserID, 2=Age, 3=ParticipantType,
    #                                4=HearingStatus, 5=SignProficiency, 6=YearsSigning,
    #                                117=StudiedSLInSchool
    participant_idx = 0
    for idx, dr in enumerate(ws1.iter_rows(min_row=2, values_only=True)):
        if not dr[1]:          # skip blank rows
            continue
        participant_idx += 1
        f_row = F_ODD if participant_idx % 2 == 0 else F_EVEN

        values = [
            participant_idx,   # sequential number
            dr[0],             # Timestamp
            dr[1],             # User ID
            dr[2],             # Age
            dr[3],             # Participant Type
            dr[4],             # Hearing Status
            dr[5],             # Sign Proficiency
            dr[6],             # Years of Signing
            dr[117],           # Studied SL in school?
        ]
        for c, v in enumerate(values, 1):
            align = ALIGN_CENTER if isinstance(v, (int, float)) else ALIGN_LEFT
            write_cell(ws, row, c, v, fill=f_row, font=FONT_NORMAL, align=align)
        ws.row_dimensions[row].height = 18
        row += 1

    ws.freeze_panes = "A3"
    col_widths = [5, 22, 13, 8, 26, 30, 20, 20, 24]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ─────────────────────────────────────────────────────────────────
#  Session Feedback tab  (Part III end-of-session questions)
# ─────────────────────────────────────────────────────────────────
def build_session_feedback_tab(wb_out, ws3):
    ws = wb_out.create_sheet(title="10. Session Feedback")
    row = 1
    title = "Post-Session Feedback  (Form Part III – end of session questions)"
    write_cell(ws, row, 1, title, fill=F_SECTION,
               font=FONT_WHITE_BOLD, align=ALIGN_LEFT)
    safe_merge(ws, row, 1, row, 6)
    ws.row_dimensions[row].height = 22
    row += 2

    headers = ["Evaluator", "How to Use System",
               "Where to Use", "Improvements Requested",
               "Overall Feedback", "Observer Notes"]
    for c, h in enumerate(headers, 1):
        write_cell(ws, row, c, h, fill=F_COL_HDR, font=FONT_BOLD, align=ALIGN_CENTER)
    ws.row_dimensions[row].height = 30
    row += 1

    # Session-level fields are at 0-based indices 141-145 (cols 142-146 in Part III)
    for idx, dr in enumerate(ws3.iter_rows(min_row=2, values_only=True)):
        if not dr[1]:
            continue
        f_row = F_ODD if idx % 2 == 0 else F_EVEN
        values = [dr[1]] + list(dr[141:146])
        for c, v in enumerate(values, 1):
            write_cell(ws, row, c, v, fill=f_row,
                       font=FONT_NORMAL, align=ALIGN_LEFT)
        ws.row_dimensions[row].height = 50
        row += 1

    ws.freeze_panes = "A3"
    for ci, w in zip(range(1, 7), [13, 32, 28, 42, 32, 32]):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ─────────────────────────────────────────────────────────────────
#  Main
# ─────────────────────────────────────────────────────────────────
def main():
    print(f"Reading  {INPUT_FILE} …")
    wb_in = openpyxl.load_workbook(INPUT_FILE)

    sheet_map = {name: wb_in[name] for name in wb_in.sheetnames}

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)   # remove default empty sheet

    print("  Building tab: 0. Participant Details …")
    build_participant_details_tab(wb_out, sheet_map["Form responses 1"])

    for section in SECTIONS:
        ws_in = sheet_map[section["sheet_key"]]
        print(f"  Building tab: {section['tab_name']} …")
        build_section_tab(wb_out, section, ws_in)

    print("  Building tab: 10. Session Feedback …")
    build_session_feedback_tab(wb_out, sheet_map["Form responses 3"])

    print(f"Saving  {OUTPUT_FILE} …")
    wb_out.save(OUTPUT_FILE)
    print("Done ✓")
    print()
    print("NOTE: Yellow cells in the 'Sentence' column of fixed sentences")
    print("      contain '[Fixed – enter sentence from form]'.")
    print("      Please fill those cells with the actual sentences from")
    print("      the Google Form videos.")


if __name__ == "__main__":
    main()
